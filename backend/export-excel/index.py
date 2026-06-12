import json
import os
import base64
import psycopg2
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import timezone

SCHEMA = os.environ.get('MAIN_DB_SCHEMA', 't_p75689129_landing_chatbot_desi')
ADMIN_EMAIL = 'el037d@gmail.com'


def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])


def style_header(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4F46E5')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 16)


def fmt(dt):
    if dt is None:
        return ''
    if hasattr(dt, 'astimezone'):
        return dt.astimezone(timezone.utc).strftime('%d.%m.%Y %H:%M')
    return str(dt)


def handler(event: dict, context) -> dict:
    """Выгрузка данных в Excel: пользователи, подписки, заявки. Только для администратора."""

    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': {'Access-Control-Allow-Origin': '*', 'Access-Control-Allow-Methods': 'GET, OPTIONS', 'Access-Control-Allow-Headers': 'Content-Type, X-Authorization'}, 'body': ''}

    raw_token = (event.get('headers') or {}).get('X-Authorization', '').replace('Bearer ', '')
    if not raw_token:
        return {'statusCode': 401, 'headers': {'Access-Control-Allow-Origin': '*'}, 'body': json.dumps({'ok': False, 'error': 'Нет токена'})}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        f"SELECT u.email FROM {SCHEMA}.sessions s JOIN {SCHEMA}.users u ON u.id=s.user_id WHERE s.token=%s AND s.expires_at > NOW()",
        (raw_token,)
    )
    row = cur.fetchone()
    if not row or row[0] != ADMIN_EMAIL:
        return {'statusCode': 403, 'headers': {'Access-Control-Allow-Origin': '*'}, 'body': json.dumps({'ok': False, 'error': 'Доступ запрещён'})}

    wb = openpyxl.Workbook()

    # Лист 1: Пользователи
    ws1 = wb.active
    ws1.title = 'Пользователи'
    headers1 = ['ID', 'Email', 'Имя', 'Дата регистрации']
    style_header(ws1, headers1)
    cur.execute(f"SELECT id, email, name, created_at FROM {SCHEMA}.users ORDER BY created_at DESC")
    for r in cur.fetchall():
        ws1.append([r[0], r[1], r[2] or '', fmt(r[3])])

    # Лист 2: Подписки / Платежи
    ws2 = wb.create_sheet('Подписки')
    headers2 = ['ID', 'Email пользователя', 'Тариф', 'Активна', 'Начало', 'Истекает']
    style_header(ws2, headers2)
    cur.execute(f"""
        SELECT s.id, u.email, s.plan, s.is_active, s.started_at, s.expires_at
        FROM {SCHEMA}.subscriptions s
        JOIN {SCHEMA}.users u ON u.id = s.user_id
        WHERE s.plan != 'free'
        ORDER BY s.started_at DESC
    """)
    for r in cur.fetchall():
        ws2.append([r[0], r[1], r[2], 'Да' if r[3] else 'Нет', fmt(r[4]), fmt(r[5])])

    # Лист 3: Контактные заявки (payment_links как заявки)
    ws3 = wb.create_sheet('Заявки')
    headers3 = ['ID', 'Email', 'Тариф', 'Использована', 'Создана', 'Истекает']
    style_header(ws3, headers3)
    cur.execute(f"SELECT id, email, plan, used, created_at, expires_at FROM {SCHEMA}.payment_links ORDER BY created_at DESC")
    for r in cur.fetchall():
        ws3.append([r[0], r[1], r[2], 'Да' if r[3] else 'Нет', fmt(r[4]), fmt(r[5])])

    cur.close()
    conn.close()

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode()

    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Origin': '*',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename="UrokAI.xlsx"',
        },
        'body': b64,
        'isBase64Encoded': True,
    }
